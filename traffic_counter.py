import sys
import json
import os
from datetime import datetime
from collections import defaultdict

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QPushButton, QLineEdit, QGroupBox, QScrollArea,
    QMessageBox, QFileDialog, QToolTip, QDialog, QCheckBox, QDialogButtonBox,
    QSizePolicy, QTableWidget, QTableWidgetItem, QHeaderView, QInputDialog,
    QButtonGroup, QRadioButton, QToolBar, QStatusBar
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QPixmap, QPainter, QColor, QIcon, QAction

# Экспорт в Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ------------------------------------------------------------
# Конфигурация направлений
ALL_DIRECTIONS = ["N", "S", "E", "W"]
DIRECTION_NAMES = {"N": "Север (N)", "S": "Юг (S)", "E": "Восток (E)", "W": "Запад (W)"}

def get_ordered_exits(entry):
    if entry == "N": return ["E", "S", "W", "N"]
    elif entry == "S": return ["W", "N", "E", "S"]
    elif entry == "E": return ["S", "W", "N", "E"]
    elif entry == "W": return ["N", "E", "S", "W"]
    else: return []

# ------------------------------------------------------------
# Класс для хранения типа ТС
class VehicleType:
    def __init__(self, name, description="", is_public=False):
        self.name = name
        self.description = description
        self.is_public = is_public

    def to_dict(self):
        return {"name": self.name, "description": self.description, "is_public": self.is_public}

    @staticmethod
    def from_dict(data):
        return VehicleType(data["name"], data.get("description", ""), data.get("is_public", False))

# ------------------------------------------------------------
# Первое окно: выбор направлений (без изменений)
class DirectionSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выбор направлений перекрёстка")
        self.setModal(True)
        self.setMinimumWidth(500)
        layout = QVBoxLayout(self)

        info = QLabel("Выберите въезды (откуда) и выезды (куда). Будут созданы все комбинации.")
        info.setWordWrap(True)
        layout.addWidget(info)

        # Въезды
        entry_group = QGroupBox("Въезды (откуда едут)")
        entry_layout = QHBoxLayout(entry_group)
        self.entry_cbs = {}
        for d in ALL_DIRECTIONS:
            cb = QCheckBox(DIRECTION_NAMES[d])
            cb.setChecked(True)
            self.entry_cbs[d] = cb
            entry_layout.addWidget(cb)
        layout.addWidget(entry_group)

        # Выезды
        exit_group = QGroupBox("Выезды (куда могут направляться)")
        exit_layout = QHBoxLayout(exit_group)
        self.exit_cbs = {}
        for d in ALL_DIRECTIONS:
            cb = QCheckBox(DIRECTION_NAMES[d])
            cb.setChecked(True)
            self.exit_cbs[d] = cb
            exit_layout.addWidget(cb)
        layout.addWidget(exit_group)

        btn_layout = QHBoxLayout()
        def all_entries(checked): [cb.setChecked(checked) for cb in self.entry_cbs.values()]
        def all_exits(checked): [cb.setChecked(checked) for cb in self.exit_cbs.values()]
        btn_layout.addWidget(QPushButton("Въезды: все", clicked=lambda: all_entries(True)))
        btn_layout.addWidget(QPushButton("Въезды: нет", clicked=lambda: all_entries(False)))
        btn_layout.addWidget(QPushButton("Выезды: все", clicked=lambda: all_exits(True)))
        btn_layout.addWidget(QPushButton("Выезды: нет", clicked=lambda: all_exits(False)))
        layout.addLayout(btn_layout)

        contact = QLabel("По вопросам обращаться к @Kango911")
        contact.setAlignment(Qt.AlignCenter)
        contact.setStyleSheet("color: gray; font-style: italic;")
        layout.addWidget(contact)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_selected(self):
        entries = [d for d, cb in self.entry_cbs.items() if cb.isChecked()]
        exits = [d for d, cb in self.exit_cbs.items() if cb.isChecked()]
        return entries, exits

# ------------------------------------------------------------
# Диалог выбора источника типов ТС (без изменений)
class LoadSourceDialog(QDialog):
    def __init__(self, auto_save_exists, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Загрузка типов транспортных средств")
        self.setModal(True)
        self.resize(400, 200)

        layout = QVBoxLayout(self)
        label = QLabel("Выберите источник загрузки списка типов ТС:")
        layout.addWidget(label)

        self.auto_radio = QRadioButton("Автосохранённый файл (предыдущие настройки)")
        self.file_radio = QRadioButton("Из внешнего JSON-файла...")
        self.default_radio = QRadioButton("Стандартный набор (по умолчанию)")

        if not auto_save_exists:
            self.auto_radio.setEnabled(False)
            self.auto_radio.setText("Автосохранённый файл (не найден)")

        layout.addWidget(self.auto_radio)
        layout.addWidget(self.file_radio)
        layout.addWidget(self.default_radio)

        self.auto_radio.setChecked(True)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def get_choice(self):
        if self.auto_radio.isChecked():
            return "auto"
        elif self.file_radio.isChecked():
            return "file"
        else:
            return "default"

# ------------------------------------------------------------
# Второе окно: редактор типов ТС (без изменений, кроме компактности?)
class VehicleTypesDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование типов транспортных средств")
        self.setModal(True)
        self.resize(750, 500)

        layout = QVBoxLayout(self)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Вкл.", "Название", "Описание", "Общественный"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        self.add_btn = QPushButton("Добавить")
        self.edit_btn = QPushButton("Редактировать")
        self.del_btn = QPushButton("Удалить")
        self.load_btn = QPushButton("Загрузить JSON")
        self.save_btn = QPushButton("Сохранить JSON")
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.edit_btn)
        btn_layout.addWidget(self.del_btn)
        btn_layout.addWidget(self.load_btn)
        btn_layout.addWidget(self.save_btn)
        layout.addLayout(btn_layout)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

        self.add_btn.clicked.connect(self.add_type)
        self.edit_btn.clicked.connect(self.edit_type)
        self.del_btn.clicked.connect(self.del_type)
        self.load_btn.clicked.connect(self.load_from_json)
        self.save_btn.clicked.connect(self.save_to_json)

        self.auto_save_path = os.path.join(os.path.dirname(sys.argv[0]), "vehicle_types_auto.json")
        self.default_types = self.get_default_types()
        self.load_initial_types()

    def get_default_types(self):
        return [
            VehicleType("car", "Легковые автомобили", False),
            VehicleType("mini_bus", "Микроавтобусы (газель, скорая)", True),
            VehicleType("middle_bus", "Средние автобусы (ПАЗ)", True),
            VehicleType("bus", "Большие автобусы (ЛиАЗ)", True),
            VehicleType("mini_truck", "Малые грузовики (до 2 т)", False),
            VehicleType("middle_truck", "Средние грузовики (2-6 т)", False),
            VehicleType("truck", "Тяжёлые грузовики (>6 т)", False),
            VehicleType("road_train", "Автопоезда", False),
            VehicleType("trol", "Троллейбусы", True),
            VehicleType("tram", "Трамваи", True)
        ]

    def load_initial_types(self):
        auto_exists = os.path.exists(self.auto_save_path)
        source_dialog = LoadSourceDialog(auto_exists, self)
        if source_dialog.exec() != QDialog.Accepted:
            self.all_types = self.default_types[:]
            return

        choice = source_dialog.get_choice()
        if choice == "auto" and auto_exists:
            self.load_auto_save()
            if not self.all_types:
                self.all_types = self.default_types[:]
        elif choice == "file":
            path, _ = QFileDialog.getOpenFileName(self, "Выберите JSON файл с типами", "", "JSON (*.json)")
            if path:
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    self.all_types = [VehicleType.from_dict(d) for d in data]
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл:\n{e}")
                    self.all_types = self.default_types[:]
            else:
                self.all_types = self.default_types[:]
        else:
            self.all_types = self.default_types[:]
        self.refresh_table()

    def refresh_table(self):
        self.table.setRowCount(len(self.all_types))
        for i, vt in enumerate(self.all_types):
            chk = QCheckBox()
            chk.setChecked(True)
            self.table.setCellWidget(i, 0, chk)
            self.table.setItem(i, 1, QTableWidgetItem(vt.name))
            self.table.setItem(i, 2, QTableWidgetItem(vt.description))
            pub_chk = QCheckBox()
            pub_chk.setChecked(vt.is_public)
            self.table.setCellWidget(i, 3, pub_chk)
        self.table.resizeRowsToContents()

    def add_type(self):
        name, ok = QInputDialog.getText(self, "Новый тип", "Название типа:")
        if ok and name.strip():
            desc, ok2 = QInputDialog.getText(self, "Описание", "Описание:", QLineEdit.Normal, "")
            desc = desc if ok2 else ""
            is_public = QMessageBox.question(self, "Общественный?", "Это общественный транспорт?",
                                             QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes
            new_type = VehicleType(name.strip(), desc, is_public)
            self.all_types.append(new_type)
            self.refresh_table()
            self.save_auto_save()

    def edit_type(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите тип для редактирования")
            return
        vt = self.all_types[row]
        name, ok = QInputDialog.getText(self, "Редактирование", "Название:", QLineEdit.Normal, vt.name)
        if ok and name.strip():
            desc, ok2 = QInputDialog.getText(self, "Описание", "Описание:", QLineEdit.Normal, vt.description)
            desc = desc if ok2 else vt.description
            is_public = QMessageBox.question(self, "Общественный?", "Это общественный транспорт?",
                                             QMessageBox.Yes | QMessageBox.No,
                                             QMessageBox.Yes if vt.is_public else QMessageBox.No) == QMessageBox.Yes
            vt.name = name.strip()
            vt.description = desc
            vt.is_public = is_public
            self.refresh_table()
            self.save_auto_save()

    def del_type(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Ошибка", "Выберите тип для удаления")
            return
        if len(self.all_types) == 1:
            QMessageBox.warning(self, "Ошибка", "Нельзя удалить единственный тип")
            return
        reply = QMessageBox.question(self, "Удаление", f"Удалить тип '{self.all_types[row].name}'?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            del self.all_types[row]
            self.refresh_table()
            self.save_auto_save()

    def save_auto_save(self):
        try:
            data = [vt.to_dict() for vt in self.all_types]
            with open(self.auto_save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка автосохранения: {e}")

    def load_auto_save(self):
        if os.path.exists(self.auto_save_path):
            try:
                with open(self.auto_save_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                loaded = [VehicleType.from_dict(d) for d in data]
                if loaded:
                    self.all_types = loaded
                    return True
            except Exception as e:
                print(f"Ошибка загрузки автосохранения: {e}")
        return False

    def save_to_json(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить типы", "vehicle_types.json", "JSON (*.json)")
        if path:
            data = [vt.to_dict() for vt in self.all_types]
            try:
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                QMessageBox.information(self, "Сохранено", f"Сохранено {len(self.all_types)} типов")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def load_from_json(self):
        path, _ = QFileDialog.getOpenFileName(self, "Загрузить типы", "", "JSON (*.json)")
        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.all_types = [VehicleType.from_dict(d) for d in data]
                self.refresh_table()
                self.save_auto_save()
                QMessageBox.information(self, "Загружено", f"Загружено {len(self.all_types)} типов")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def get_selected_types(self):
        selected = []
        for i in range(self.table.rowCount()):
            chk = self.table.cellWidget(i, 0)
            if chk and chk.isChecked():
                vt = self.all_types[i]
                pub_chk = self.table.cellWidget(i, 3)
                if pub_chk:
                    vt.is_public = pub_chk.isChecked()
                selected.append(vt)
        return selected

    def accept(self):
        self.save_auto_save()
        super().accept()

# ------------------------------------------------------------
# Компактное главное окно с возможностью сворачивания групп
class CollapsibleGroupBox(QGroupBox):
    def __init__(self, title, parent=None):
        super().__init__(title, parent)
        self.setCheckable(True)
        self.setChecked(True)
        self.toggled.connect(self.on_toggled)
        self.setStyleSheet("""
            QGroupBox::indicator {
                subcontrol-position: top left;
                width: 16px;
                height: 16px;
            }
            QGroupBox {
                margin-top: 12px;
            }
        """)

    def on_toggled(self, checked):
        for child in self.findChildren(QWidget):
            if child != self:
                child.setVisible(checked)

class TrafficCounterApp(QMainWindow):
    def __init__(self, entries, exits, vehicle_types):
        super().__init__()
        self.entries = entries
        self.exits = exits
        self.vehicle_types = vehicle_types

        # Генерация направлений
        self.directions = {}
        for entry in self.entries:
            ordered = get_ordered_exits(entry)
            for ex in ordered:
                if ex in self.exits:
                    code = entry + ex
                    self.directions[code] = f"{entry} → {ex}"

        if not self.directions:
            QMessageBox.critical(self, "Ошибка", "Нет направлений для выбранных въездов/выездов")
            sys.exit(1)

        self.setWindowTitle("Счётчик транспортных средств на перекрёстке")
        self.setMinimumSize(500, 400)   # компактный минимальный размер
        self.setWindowIcon(self.create_k9_icon())

        self.counters = defaultdict(int)
        self.buttons = {}

        # Центральный виджет с прокруткой
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(2, 2, 2, 2)
        main_layout.setSpacing(2)

        # Верхняя панель - компактная
        top_widget = QWidget()
        top_layout = QHBoxLayout(top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)
        self.cross_name = QLineEdit()
        self.cross_name.setPlaceholderText("Перекрёсток")
        self.cross_name.setText("Перекрёсток")
        self.cross_name.setMaximumWidth(200)
        top_layout.addWidget(QLabel("Перекрёсток:"))
        top_layout.addWidget(self.cross_name)

        self.date_edit = QLineEdit()
        self.date_edit.setText(datetime.now().strftime("%Y-%m-%d %H:%M"))
        self.date_edit.setMaximumWidth(150)
        top_layout.addWidget(QLabel("Дата:"))
        top_layout.addWidget(self.date_edit)

        self.export_btn = QPushButton("Excel")
        self.export_btn.setFixedWidth(60)
        self.export_btn.clicked.connect(self.export_to_excel)
        top_layout.addWidget(self.export_btn)

        top_layout.addStretch()
        main_layout.addWidget(top_widget)

        # Прокручиваемая область с группами
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setSpacing(5)
        container_layout.setContentsMargins(2, 2, 2, 2)

        for entry in self.entries:
            group = CollapsibleGroupBox(f"Направления из {DIRECTION_NAMES[entry]}")
            group_layout = QGridLayout(group)
            group_layout.setVerticalSpacing(2)
            group_layout.setHorizontalSpacing(2)

            # Заголовки типов (только названия, без "i" для компактности? но оставим i)
            for col, vt in enumerate(self.vehicle_types):
                widget = QWidget()
                h_layout = QHBoxLayout(widget)
                h_layout.setContentsMargins(0,0,0,0)
                label = QLabel(vt.name)
                label.setAlignment(Qt.AlignCenter)
                label.setWordWrap(False)
                label.setFixedHeight(20)
                info_btn = QPushButton("i")
                info_btn.setFixedSize(18,18)
                info_btn.setToolTip(vt.description if vt.description else "Нет описания")
                h_layout.addWidget(label, 1)
                h_layout.addWidget(info_btn, 0)
                group_layout.addWidget(widget, 0, col+1)
                group_layout.setColumnStretch(col+1, 1)

            # Строки направлений
            ordered = get_ordered_exits(entry)
            for i, ex in enumerate(ordered):
                if ex not in self.exits:
                    continue
                dir_code = entry + ex
                dir_label = QLabel(self.directions[dir_code])
                dir_label.setStyleSheet("font-weight: bold;")
                dir_label.setFixedHeight(30)
                group_layout.addWidget(dir_label, i+1, 0)

                for col, vt in enumerate(self.vehicle_types):
                    key = (dir_code, vt.name)
                    btn = QPushButton(str(self.counters[key]))
                    btn.setFixedSize(50, 30)   # компактные кнопки
                    btn.setStyleSheet("""
                        QPushButton {
                            background-color: #e0e0e0;
                            border: 1px solid #aaa;
                            border-radius: 3px;
                            font-size: 11px;
                            font-weight: bold;
                        }
                        QPushButton:hover {
                            background-color: #c0c0c0;
                        }
                    """)
                    btn.clicked.connect(lambda checked, d=dir_code, t=vt.name: self.inc(d, t))
                    btn.setContextMenuPolicy(Qt.CustomContextMenu)
                    btn.customContextMenuRequested.connect(lambda pos, d=dir_code, t=vt.name: self.dec(d, t))
                    self.buttons[key] = btn
                    group_layout.addWidget(btn, i+1, col+1)

            container_layout.addWidget(group)

        scroll.setWidget(container)
        main_layout.addWidget(scroll)

        # Статус бар
        self.statusBar().showMessage("Левая кнопка +1, правая -1 | Группы можно сворачивать")

    def create_k9_icon(self):
        pixmap = QPixmap(64,64)
        pixmap.fill(QColor(42,130,218))
        painter = QPainter(pixmap)
        painter.setPen(QColor(255,255,255))
        font = QFont("Arial", 24, QFont.Bold)
        painter.setFont(font)
        painter.drawText(pixmap.rect(), Qt.AlignCenter, "K9")
        painter.end()
        return QIcon(pixmap)

    def inc(self, direction, vtype):
        key = (direction, vtype)
        self.counters[key] += 1
        self.buttons[key].setText(str(self.counters[key]))

    def dec(self, direction, vtype):
        key = (direction, vtype)
        if self.counters[key] > 0:
            self.counters[key] -= 1
            self.buttons[key].setText(str(self.counters[key]))

    def export_to_excel(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Ошибка", "Установите openpyxl: pip install openpyxl")
            return
        cross = self.cross_name.text().strip() or "Без названия"
        date_str = self.date_edit.text().strip() or datetime.now().strftime("%Y-%m-%d %H:%M")
        filename = f"{cross}_{date_str.replace(' ', '_').replace(':', '-')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить Excel", filename, "*.xlsx")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Перекрёсток"

        ws.merge_cells('A1:I1')
        ws['A1'] = f"Перекрёсток: {cross}  |  Дата: {date_str}"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ["Направление"] + [vt.name for vt in self.vehicle_types]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = h
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[chr(64+col)].width = 20 if col == 1 else 15

        row = 4
        direction_data = {}
        for entry in self.entries:
            ordered = get_ordered_exits(entry)
            for ex in ordered:
                if ex not in self.exits:
                    continue
                dir_code = entry + ex
                dir_display = self.directions[dir_code]
                ws.cell(row=row, column=1, value=dir_display)
                direction_data[dir_code] = {}
                for col, vt in enumerate(self.vehicle_types, 2):
                    cnt = self.counters[(dir_code, vt.name)]
                    ws.cell(row=row, column=col, value=cnt)
                    direction_data[dir_code][vt.name] = cnt
                row += 1

        total_all = 0
        total_no_public = 0
        for types in direction_data.values():
            for vname, cnt in types.items():
                total_all += cnt
                vt = next((vt for vt in self.vehicle_types if vt.name == vname), None)
                if vt and not vt.is_public:
                    total_no_public += cnt
        percent_no_public = (total_no_public / total_all * 100) if total_all else 0

        row += 1
        ws.cell(row=row, column=1, value="Общее количество ТС:")
        ws.cell(row=row, column=2, value=total_all)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Количество без общественного:")
        ws.cell(row=row, column=2, value=total_no_public)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Доля без общественного (%):")
        ws.cell(row=row, column=2, value=f"{percent_no_public:.2f}%")
        row += 2

        ws.cell(row=row, column=1, value="ДОЛИ ПОВОРОТОВ ПО ВЪЕЗДАМ")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        entry_totals = defaultdict(int)
        entry_exit_counts = defaultdict(lambda: defaultdict(int))
        for dir_code, types in direction_data.items():
            entry = dir_code[0]
            exit_ = dir_code[1]
            total_dir = sum(types.values())
            entry_totals[entry] += total_dir
            entry_exit_counts[entry][exit_] += total_dir

        for entry in self.entries:
            total_entry = entry_totals.get(entry, 0)
            ws.cell(row=row, column=1, value=f"Въезд: {DIRECTION_NAMES[entry]}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            ordered = get_ordered_exits(entry)
            filtered = [ex for ex in ordered if ex in self.exits]
            for col, ex in enumerate(filtered, 2):
                ws.cell(row=row, column=col, value=f"→ {ex}")
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
            for ex in filtered:
                cnt = entry_exit_counts[entry].get(ex, 0)
                percent = (cnt / total_entry * 100) if total_entry else 0
                col = filtered.index(ex) + 2
                ws.cell(row=row, column=col, value=f"{percent:.1f}%")
            row += 2

        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for r in range(3, row):
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    cell.border = thin

        wb.save(path)
        QMessageBox.information(self, "Готово", f"Экспорт завершён:\n{path}")

# ------------------------------------------------------------
def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    dir_dialog = DirectionSelectionDialog()
    if dir_dialog.exec() != QDialog.Accepted:
        sys.exit(0)
    entries, exits = dir_dialog.get_selected()
    if not entries or not exits:
        QMessageBox.critical(None, "Ошибка", "Не выбраны въезды или выезды")
        sys.exit(1)

    types_dialog = VehicleTypesDialog()
    if types_dialog.exec() != QDialog.Accepted:
        sys.exit(0)
    vehicle_types = types_dialog.get_selected_types()
    if not vehicle_types:
        QMessageBox.critical(None, "Ошибка", "Не выбран ни один тип ТС")
        sys.exit(1)

    window = TrafficCounterApp(entries, exits, vehicle_types)
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()